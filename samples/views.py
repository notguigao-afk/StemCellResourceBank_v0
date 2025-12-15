from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
from django.db.models import Q, Count
from django.http import HttpResponse
from django.utils.translation import gettext as _
from django.utils import timezone
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import Sample, SiteSettings
from .forms import SampleForm, SiteSettingsForm


# Permission checking functions
def is_staff_or_admin(user):
    """Check if user is staff member or admin"""
    return user.is_authenticated and (user.groups.filter(name='Lab Staff').exists() or user.is_superuser)


def is_admin(user):
    """Check if user is admin"""
    return user.is_authenticated and user.is_superuser


# Authentication views
def login_view(request):
    """Login view - main entry point"""
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, _('Welcome back, %(username)s!') % {'username': username})
                next_url = request.GET.get('next', 'home')
                return redirect(next_url)
        else:
            messages.error(request, _('Invalid username or password.'))
    else:
        form = AuthenticationForm()
    
    # Get site settings for login page
    site_settings = SiteSettings.get_settings()
    
    return render(request, 'samples/login.html', {
        'form': form,
        'site_settings': site_settings,
    })


def logout_view(request):
    """Logout view"""
    logout(request)
    messages.info(request, _('You have been logged out successfully.'))
    return redirect('login')


# Dashboard - Home view (requires login)
@login_required
def home(request):
    """Dashboard home page - shows summary statistics"""
    today = timezone.now().date()
    seven_days_ago = today - timedelta(days=7)
    thirty_days_ago = today - timedelta(days=30)
    
    # Get statistics
    total_samples = Sample.objects.count()
    available_samples = Sample.objects.filter(status='AVAILABLE').count()
    in_use_samples = Sample.objects.filter(status='IN_USE').count()
    depleted_samples = Sample.objects.filter(status='DEPLETED').count()
    reserved_samples = Sample.objects.filter(status='RESERVED').count()
    quarantine_samples = Sample.objects.filter(status='QUARANTINE').count()
    
    # Samples by type
    samples_by_type = Sample.objects.values('sample_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Recently modified samples (last 7 days)
    recent_samples = Sample.objects.filter(
        updated_at__gte=seven_days_ago
    ).order_by('-updated_at')[:10]
    
    # Samples expiring soon (next 30 days)
    expiring_soon = Sample.objects.filter(
        expiration_date__gte=today,
        expiration_date__lte=today + timedelta(days=30)
    ).order_by('expiration_date')[:5]
    
    # Low stock samples (quantity < 3)
    low_stock = Sample.objects.filter(
        status='AVAILABLE',
        quantity__lt=3,
        quantity__gt=0
    ).order_by('quantity')[:5]
    
    context = {
        'today': today,
        'total_samples': total_samples,
        'available_samples': available_samples,
        'in_use_samples': in_use_samples,
        'depleted_samples': depleted_samples,
        'reserved_samples': reserved_samples,
        'quarantine_samples': quarantine_samples,
        'samples_by_type': samples_by_type,
        'recent_samples': recent_samples,
        'expiring_soon': expiring_soon,
        'low_stock': low_stock,
        'sample_types': dict(Sample.SAMPLE_TYPE_CHOICES),
    }
    return render(request, 'samples/home.html', context)


# Sample management views - require authentication
@login_required
@user_passes_test(is_staff_or_admin, login_url='login')
def sample_list(request):
    """List all samples - for lab staff and admins"""
    samples = Sample.objects.select_related('created_by').all()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        samples = samples.filter(
            Q(sample_id__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(sample_type__icontains=search_query) |
            Q(storage_location__icontains=search_query)
        )
    
    # Filter by type
    sample_type = request.GET.get('type', '')
    if sample_type:
        samples = samples.filter(sample_type=sample_type)
    
    # Filter by status
    status = request.GET.get('status', '')
    if status:
        samples = samples.filter(status=status)
    
    context = {
        'samples': samples,
        'search_query': search_query,
        'selected_type': sample_type,
        'selected_status': status,
        'sample_types': Sample.SAMPLE_TYPE_CHOICES,
        'status_choices': Sample.STATUS_CHOICES,
    }
    return render(request, 'samples/sample_list.html', context)


@login_required
@user_passes_test(is_staff_or_admin, login_url='login')
def sample_detail(request, pk):
    """View sample details - for lab staff and admins"""
    sample = get_object_or_404(Sample, pk=pk)
    
    # Get history for this sample
    history = sample.history.all()[:20]  # Last 20 changes
    
    return render(request, 'samples/sample_detail.html', {
        'sample': sample,
        'history': history,
    })


@login_required
@user_passes_test(is_staff_or_admin, login_url='login')
def sample_create(request):
    """Create a new sample - for lab staff and admins"""
    if request.method == 'POST':
        form = SampleForm(request.POST, request.FILES)
        if form.is_valid():
            sample = form.save(commit=False)
            sample.created_by = request.user
            sample.save()
            messages.success(request, _('Sample %(sample_id)s created successfully!') % {'sample_id': sample.sample_id})
            return redirect('sample_detail', pk=sample.pk)
    else:
        form = SampleForm()
    
    return render(request, 'samples/sample_form.html', {
        'form': form,
        'title': _('Add New Sample'),
        'button_text': _('Create Sample')
    })


@login_required
@user_passes_test(is_staff_or_admin, login_url='login')
def sample_update(request, pk):
    """Update an existing sample - for lab staff and admins"""
    sample = get_object_or_404(Sample, pk=pk)
    
    if request.method == 'POST':
        form = SampleForm(request.POST, request.FILES, instance=sample)
        if form.is_valid():
            form.save()
            messages.success(request, _('Sample %(sample_id)s updated successfully!') % {'sample_id': sample.sample_id})
            return redirect('sample_detail', pk=sample.pk)
    else:
        form = SampleForm(instance=sample)
    
    return render(request, 'samples/sample_form.html', {
        'form': form,
        'sample': sample,
        'title': _('Edit Sample: %(sample_id)s') % {'sample_id': sample.sample_id},
        'button_text': _('Update Sample')
    })


@login_required
@user_passes_test(is_staff_or_admin, login_url='login')
def sample_delete(request, pk):
    """Delete a sample - for lab staff and admins"""
    sample = get_object_or_404(Sample, pk=pk)
    
    if request.method == 'POST':
        sample_id = sample.sample_id
        sample.delete()
        messages.success(request, _('Sample %(sample_id)s deleted successfully!') % {'sample_id': sample_id})
        return redirect('sample_list')
    
    return render(request, 'samples/sample_confirm_delete.html', {'sample': sample})


# Export functionality
@login_required
@user_passes_test(is_staff_or_admin, login_url='login')
def export_samples(request):
    """Export samples to Excel file"""
    # Get selected sample IDs and columns from request
    sample_ids = request.GET.getlist('samples')
    columns = request.GET.getlist('columns')
    
    # Default columns if none selected
    if not columns:
        columns = ['sample_id', 'name', 'sample_type', 'status', 'quantity', 
                   'storage_location', 'viability', 'collection_date', 'expiration_date']
    
    # Get samples (filtered if IDs provided)
    if sample_ids:
        samples = Sample.objects.filter(pk__in=sample_ids)
    else:
        # Apply same filters as list view
        samples = Sample.objects.all()
        
        search_query = request.GET.get('search', '')
        if search_query:
            samples = samples.filter(
                Q(sample_id__icontains=search_query) |
                Q(name__icontains=search_query) |
                Q(description__icontains=search_query)
            )
        
        sample_type = request.GET.get('type', '')
        if sample_type:
            samples = samples.filter(sample_type=sample_type)
        
        status = request.GET.get('status', '')
        if status:
            samples = samples.filter(status=status)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Samples"
    
    # Column definitions with display names
    column_names = {
        'sample_id': _('Sample ID'),
        'name': _('Sample Name'),
        'sample_type': _('Sample Type'),
        'status': _('Status'),
        'quantity': _('Quantity'),
        'passage_number': _('Passage Number'),
        'storage_location': _('Storage Location'),
        'source': _('Source'),
        'donor_info': _('Donor Information'),
        'description': _('Description'),
        'viability': _('Viability (%)'),
        'collection_date': _('Collection Date'),
        'storage_date': _('Storage Date'),
        'expiration_date': _('Expiration Date'),
        'quality_control_notes': _('QC Notes'),
        'research_use_only': _('Research Use Only'),
        'created_by': _('Created By'),
        'created_at': _('Created At'),
        'updated_at': _('Updated At'),
    }
    
    # Style definitions
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, col_key in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(column_names.get(col_key, col_key)))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, sample in enumerate(samples, 2):
        for col_idx, col_key in enumerate(columns, 1):
            value = getattr(sample, col_key, '')
            
            # Handle special fields
            if col_key == 'sample_type':
                value = sample.get_sample_type_display()
            elif col_key == 'status':
                value = sample.get_status_display()
            elif col_key == 'created_by':
                value = sample.created_by.username if sample.created_by else ''
            elif col_key == 'research_use_only':
                value = _('Yes') if value else _('No')
            elif hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d')
            elif hasattr(value, 'isoformat'):
                value = value.isoformat()
            
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
            cell.border = thin_border
    
    # Adjust column widths
    for col_idx, col_key in enumerate(columns, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=samples_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


# Site settings view (admin only)
@login_required
@user_passes_test(is_admin, login_url='login')
def site_settings_view(request):
    """Edit site settings including logo"""
    settings = SiteSettings.get_settings()
    
    if request.method == 'POST':
        form = SiteSettingsForm(request.POST, request.FILES, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, _('Site settings updated successfully!'))
            return redirect('site_settings')
    else:
        form = SiteSettingsForm(instance=settings)
    
    return render(request, 'samples/site_settings.html', {
        'form': form,
        'settings': settings,
    })
